from __future__ import annotations

import json
from pathlib import Path
from typing import Dict, List, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image as PILImage

from .sample_requests import load_enriched_selections


HEADERS = [
    "图片",
    "商品ID",
    "供应商",
    "类目",
    "选中款/颜色",
    "款号",
    "颜色",
    "尺码",
    "面料/材质",
    "价格",
    "库存/起订量",
    "发货/排单",
    "状态",
    "供应商原话",
    "供应商补充图片",
]

STATUS_LABELS = {
    "product_info_received": "信息已回",
    "product_info_received_stock_pending": "信息已回，库存待复核",
    "waiting_reply": "待回复",
    "waiting_product_info": "待商品信息",
    "supplier_no_reply": "供应商未回复",
    "needs_mapping_review": "资料需人工匹配",
}


def build_ops_table(
    selection_path: Path,
    supplier_replies_path: Path,
    output_path: Path,
    root_dir: Path | None = None,
    title: str = "选款商品信息表",
) -> Path:
    root = root_dir or Path.cwd()
    selections = load_enriched_selections(selection_path)
    replies = _load_replies(supplier_replies_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    thumb_dir = output_path.parent / f"{output_path.stem}_assets"
    thumb_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "选款商品信息"
    _style_title(ws, title)
    ws.append([])
    ws.append(HEADERS)
    _style_header(ws, 4)

    for row_index, selection in enumerate(_sort_selections(selections), start=5):
        reply = replies.get(selection["product_id"], {})
        _write_row(ws, row_index, selection, reply)
        image_path = root / selection["primary_image"]
        if image_path.exists():
            thumb_path = thumb_dir / f"{selection['product_id']}.jpg"
            make_thumbnail(image_path, thumb_path)
            image = XLImage(str(thumb_path))
            image.width = 132
            image.height = 132
            ws.add_image(image, f"A{row_index}")
        info_image_path = _first_existing_info_image(reply, root)
        if info_image_path:
            thumb_path = thumb_dir / f"{selection['product_id']}_info.jpg"
            make_thumbnail(info_image_path, thumb_path)
            image = XLImage(str(thumb_path))
            image.width = 132
            image.height = 132
            ws.add_image(image, f"O{row_index}")
        ws.row_dimensions[row_index].height = 104

    _set_dimensions(ws)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:O{4 + len(selections)}"
    wb.save(output_path)
    _verify_workbook(output_path, expected_rows=len(selections))
    return output_path


def make_thumbnail(src: Path, dst: Path, size=(220, 220)) -> None:
    with PILImage.open(src) as img:
        img.thumbnail(size)
        canvas = PILImage.new("RGB", size, "white")
        canvas.paste(img.convert("RGB"), ((size[0] - img.width) // 2, (size[1] - img.height) // 2))
        canvas.save(dst, quality=92)


def _load_replies(path: Path) -> Dict[str, dict]:
    if not path.exists():
        return {}
    payload = json.loads(path.read_text(encoding="utf-8"))
    items = payload["items"] if isinstance(payload, dict) else payload
    return {item["product_id"]: item for item in items}


def _sort_selections(selections: Sequence[dict]) -> List[dict]:
    category_rank = {"上衣/T恤": 1, "下装/半裙": 2, "下装/裤子": 3, "鞋履/单鞋": 4}
    return sorted(selections, key=lambda item: (category_rank.get(item.get("category"), 99), item["supplier_name"], item["product_id"]))


def _style_title(ws, title: str) -> None:
    ws.merge_cells("A1:O1")
    ws["A1"] = title
    ws["A1"].font = Font(name="PingFang SC", bold=True, size=18, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:O2")
    ws["A2"] = "来源：选款结果 + 供应商企业微信回复"
    ws["A2"].font = Font(name="PingFang SC", size=10, color="5B6770")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")


def _style_header(ws, header_row: int) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = Font(name="PingFang SC", bold=True, color="1F2937")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()


def _write_row(ws, row_index: int, selection: dict, reply: dict) -> None:
    values = [
        "",
        selection["product_id"],
        selection["supplier_name"],
        selection["category"],
        selection.get("selected_variant", ""),
        reply.get("style_no", ""),
        reply.get("color") or "",
        reply.get("sizes") or "",
        reply.get("material") or "",
        reply.get("price") or "",
        reply.get("stock_or_moq") or "",
        reply.get("lead_time") or "",
        STATUS_LABELS.get(reply.get("status"), reply.get("status") or "待商品信息"),
        reply.get("raw_reply") or "",
        "见图" if reply.get("info_image_paths") else "",
    ]
    fill = PatternFill("solid", fgColor="FFF8E5") if "待" in values[12] else PatternFill("solid", fgColor="F6FBF7")
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row_index, column=col, value=value)
        cell.font = Font(name="PingFang SC", size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = _border()
        cell.fill = fill


def _set_dimensions(ws) -> None:
    widths = {
        "A": 20,
        "B": 28,
        "C": 14,
        "D": 14,
        "E": 18,
        "F": 12,
        "G": 22,
        "H": 12,
        "I": 18,
        "J": 10,
        "K": 28,
        "L": 18,
        "M": 24,
        "N": 44,
        "O": 20,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _verify_workbook(path: Path, expected_rows: int) -> None:
    workbook = load_workbook(path)
    sheet = workbook["选款商品信息"]
    if sheet.max_row < expected_rows + 4:
        raise RuntimeError(f"ops table row count mismatch: {sheet.max_row}")
    if len(sheet._images) < expected_rows:
        raise RuntimeError(f"ops table image count mismatch: {len(sheet._images)}")


def _border() -> Border:
    thin = Side(style="thin", color="D9DEE7")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _first_existing_info_image(reply: dict, root: Path) -> Path | None:
    for raw_path in reply.get("info_image_paths", []) or []:
        path = Path(raw_path)
        candidates = [path] if path.is_absolute() else [root / path, path]
        for candidate in candidates:
            if candidate.exists():
                return candidate
    return None
